import openpyxl
import glob
from openpyxl.reader.excel import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.utils import get_column_letter
import numpy

def readDataFromXlsx(file_name, sheet_name,XLSX_PATH=""):
    non_hangup_datas = []
    hangup_datas = []
    hangup2_datas = []

    wb = load_workbook(XLSX_PATH+ file_name)
    try:
        sheet = wb[ sheet_name ]
        for row in range(2, sheet.max_row + 1):
            non_hangup_mu = sheet["A%d" % row].value
            non_hangup_coh = sheet["B%d" % row].value
            hangup_mu = sheet["C%d" % row].value
            hangup_coh = sheet["D%d" % row].value
            hangup2_mu = sheet["E%d" % row].value
            hangup2_coh = sheet["F%d" % row].value

            if not non_hangup_mu == None and not non_hangup_coh == None:
                non_hangup_datas.append((non_hangup_mu, non_hangup_coh))
            if not hangup_mu == None and not hangup_coh == None:
                hangup_datas.append((hangup_mu, hangup_coh))
            if not hangup2_mu == None and not hangup2_coh == None:
                hangup2_datas.append((hangup2_mu, hangup2_coh))
        return [non_hangup_datas, hangup_datas, hangup2_datas]
    except KeyError as e:
        print(e)
        return None